#!/usr/bin/env python3
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
from workflows.payroll_workflow import process_payroll_conversion
from workflows.email_service import send_simple_email, send_email
from workflows.excel_translate_workflow import process_excel_and_translate
from workflows.essensgeld_workflow import convert_essensgeld_from_upload
from workflows.pfleger_workflow import convert_pfleger_from_upload
from fastapi.responses import FileResponse
import os

app = FastAPI(
    title="Lohnabrechnung Konverter",
    description="API für Excel zu CSV Konvertierung mit E-Mail-Versand",
    version="1.0.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    """API Info"""
    return {
        "message": "Lohnabrechnung Konverter API",
        "version": "1.0.0",
        "endpoints": {
            "convert": "/convert",
            "test_email": "/test-email",
            "health": "/health",
            "docs": "/docs"
        },
        "status": "ready"
    }

@app.get("/health")
async def health_check():
    """Health Check für Monitoring"""
    return {"status": "healthy", "service": "lohnabrechnung-konverter"}

@app.post("/convert")
async def convert_payroll(
    file: UploadFile = File(..., description="Excel-Datei"),
    email: str = Form(..., description="E-Mail-Adresse"),
    mandant: str = Form("10001", description="Mandant"),
    abrechnungsmonat: Optional[str] = Form(None, description="Abrechnungsmonat YYYYMM"),
    sheet_name: str = Form("Tabelle1", description="Arbeitsblatt")
):
    """Excel zu CSV konvertieren und per E-Mail senden"""
    return await process_payroll_conversion(file, email, mandant, abrechnungsmonat, sheet_name)

@app.post("/test-email")
async def test_email(
    email: str = Form(..., description="Test E-Mail-Adresse")
):
    """E-Mail-Konfiguration testen"""
    try:
        success = send_simple_email(
            recipient_email=email,
            subject="Test E-Mail - Lohnabrechnung Service",
            body="Dies ist eine Test-E-Mail. Wenn Sie diese erhalten, funktioniert die E-Mail-Konfiguration."
        )
        
        if success:
            return {"message": f"Test-E-Mail erfolgreich an {email} gesendet"}
        else:
            raise HTTPException(
                status_code=500,
                detail="E-Mail konnte nicht gesendet werden. Überprüfen Sie die Konfiguration."
            )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"E-Mail-Test fehlgeschlagen: {str(e)}")

@app.post("/translate-excel")
async def translate_excel(
    file: UploadFile = File(..., description="Excel-Datei mit japanischem Text"),
    email: str = Form(..., description="E-Mail-Adresse")
):
    """Excel-Datei hochladen, japanische Texte nach Deutsch übersetzen und per E-Mail senden."""
    try:
        output_path = process_excel_and_translate(file)
        # Zeilen zählen (ohne Header)
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        rows_count = ws.max_row - 1 if ws.max_row > 1 else 0
        # Abrechnungsmonat ist hier nicht relevant, aber Pflicht für send_email
        success = send_email(email, output_path, "", rows_count)
        if success:
            return {"message": f"Übersetzte Excel-Datei erfolgreich an {email} gesendet."}
        else:
            raise HTTPException(status_code=500, detail="E-Mail-Versand fehlgeschlagen.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fehler bei der Übersetzung: {str(e)}")

@app.post("/convert-essensgeld")
async def convert_essensgeld(
    file: UploadFile = File(..., description="Essensgeld Excel-Datei"),
    email: str = Form(..., description="E-Mail-Adresse"),
    mandant: str = Form("10001", description="Mandant"),
    abrechnungsmonat: str = Form(None, description="Abrechnungsmonat YYYYMM")
):
    """Essensgeld-Excel konvertieren und per E-Mail senden."""
    try:
        output_path = convert_essensgeld_from_upload(file, mandant, abrechnungsmonat)
        rows_count = sum(1 for _ in open(output_path, encoding="utf-8")) - 1  # Zeilen ohne Header
        success = send_email(email, output_path, abrechnungsmonat or "", rows_count)
        if success:
            return {"message": f"Essensgeld erfolgreich an {email} gesendet."}
        else:
            raise HTTPException(status_code=500, detail="E-Mail-Versand fehlgeschlagen.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fehler bei der Essensgeld-Konvertierung: {str(e)}")

@app.post("/convert-pfleger")
async def convert_pfleger(
    file: UploadFile = File(..., description="Pflegeheim Excel-Datei"),
    email: str = Form(..., description="E-Mail-Adresse"),
    mandant: str = Form("10001", description="Mandant"),
    abrechnungsmonat: str = Form(None, description="Abrechnungsmonat YYYYMM")
):
    """Pflegeheim-Excel konvertieren und per E-Mail senden."""
    try:
        output_path = convert_pfleger_from_upload(file, mandant, abrechnungsmonat)
        rows_count = sum(1 for _ in open(output_path, encoding="utf-8")) - 1
        success = send_email(email, output_path, abrechnungsmonat or "", rows_count)
        if success:
            return {"message": f"Pflegeheim-Datei erfolgreich an {email} gesendet."}
        else:
            raise HTTPException(status_code=500, detail="E-Mail-Versand fehlgeschlagen.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fehler bei der Pflegeheim-Konvertierung: {str(e)}")

@app.get("/download/{filename}")
async def download_translated_excel(filename: str):
    """Download einer übersetzten Excel-Datei aus dem temporären Verzeichnis."""
    temp_dir = "/tmp"  # Standard für NamedTemporaryFile
    file_path = os.path.join(temp_dir, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Datei nicht gefunden")
    return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
