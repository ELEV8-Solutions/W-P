import os
import pandas as pd
import openpyxl
import re
from fastapi import UploadFile
from tempfile import NamedTemporaryFile

try:
    from deep_translator import GoogleTranslator
    TRANSLATOR_AVAILABLE = True
except ImportError:
    TRANSLATOR_AVAILABLE = False

def has_japanese_characters(text):
    if not text or not isinstance(text, str):
        return False
    japanese_pattern = r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FAF]'
    return bool(re.search(japanese_pattern, text))

def translate_texts(texts_to_translate):
    if not texts_to_translate or not TRANSLATOR_AVAILABLE:
        return {}
    translation_map = {}
    try:
        translator = GoogleTranslator(source='ja', target='de')
        translated_texts = translator.translate_batch(texts_to_translate)
        for original, translated in zip(texts_to_translate, translated_texts):
            if translated and translated.lower() != original.lower():
                translation_map[original] = translated
    except Exception:
        for text in texts_to_translate:
            try:
                translator = GoogleTranslator(source='ja', target='de')
                translated = translator.translate(text)
                if translated and translated.lower() != text.lower():
                    translation_map[text] = translated
            except Exception:
                pass
    return translation_map

def process_excel_and_translate(file: UploadFile) -> str:
    """
    Nimmt eine Excel-Datei, übersetzt japanische Texte nach Deutsch und gibt den Pfad zur neuen Datei zurück.
    """
    # Speichere Upload temporär
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    wb = openpyxl.load_workbook(tmp_path)
    ws = wb.active

    # Sammle alle einzigartigen japanischen Texte
    unique_japanese_texts = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and has_japanese_characters(str(cell.value)):
                unique_japanese_texts.add(str(cell.value))

    translation_map = translate_texts(list(unique_japanese_texts)) if unique_japanese_texts else {}

    # Ersetze die Werte durch Übersetzungen
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value) in translation_map:
                cell.value = translation_map[str(cell.value)]

    # Speichere die übersetzte Datei
    output_path = tmp_path.replace('.xlsx', '_translated.xlsx')
    wb.save(output_path)
    return output_path
